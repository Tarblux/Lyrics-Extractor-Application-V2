# import lyricsgenius
# import csv
# import re

# Genius API key

# genius = lyricsgenius.Genius('xxxxxxxxxxx')

# # Search for the song and get the lyrics
# song = genius.search_song('Artist', 'Song')

# pattern = r'\[.*?\]' # pattern to match anything in square brackets

# # remove the pattern from the lyrics
# result = re.sub(pattern, '', song.lyrics)

# # Create a CSV file and write the lyrics to it
# with open('lyrics.csv', 'w', newline='', encoding='utf-8') as csvfile:
#     writer = csv.writer(csvfile)
#     # write the lyrics to the CSV file 
#     writer.writerow(['Lyrics'])
#     writer.writerow(['\n'.join(result.split('\n')[1:])])


import lyricsgenius
import csv
import re
import openpyxl

#  Genius API key
genius = lyricsgenius.Genius('xxxxxxxxxxx')



# The regex pattern used to remove square brackets and their contents
pattern = r'\[.*?\]'

# CSV file and reads the artist and song names
with open('song_artist.csv', 'r', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile)
    # skips the header
    next(reader)
    # Creates a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    # Set the column headers
    worksheet.cell(row=1, column=1, value='Artist')
    worksheet.cell(row=1, column=2, value='Song')
    worksheet.cell(row=1, column=3, value='Lyrics')
    # Iterates over each row in the CSV file
    for row_num, row in enumerate(reader, start=2):
        # Searches for the song and get the lyrics
        song = genius.search_song(row[0], row[1])
        result = re.sub(pattern, '', song.lyrics)
        # Writes the artist, song, and lyrics to the worksheet
        worksheet.cell(row=row_num, column=1, value=row[0])
        worksheet.cell(row=row_num, column=2, value=row[1])
        worksheet.cell(row=row_num, column=3, value='\n'.join(result.split('\n')[1:]))
    # Saves the workbook as an xlsx file
    workbook.save('lyrics.xlsx')












